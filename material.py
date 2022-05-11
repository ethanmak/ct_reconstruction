import numpy as np
from openpyxl import load_workbook


class Material(object):
	def __init__(self):
		"""Material holds material, mev, and coeff information
		which is loaded from an xlsx spreadsheet on initialisation"""

		# set up file name, sheet name, and important header names
		filename = 'mass_attenuation_coeffs.xlsx'
		sheetname = 'Materials'
		mevname = 'MeV'

		# open workbook
		book = load_workbook(filename, read_only=True, data_only=True)

		# check for existing sheet name
		if sheetname not in book.sheetnames:
			raise IndexError(filename + ' does not contain a ' + sheetname + ' sheet')

		# load header row, containing material names
		sheet = book[sheetname]
		header = []
		for row in sheet.iter_rows(min_row=1, max_row=1):
			for cell in row:
				header.append(cell.value)

		# check first header is energy
		if mevname not in header[0]:
			raise IndexError(sheetname + ' does not contain a ' + mevname + ' header')
		
		# load the first column, which is energy values
		self.name = header[1:]
		mev = []
		for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
			for cell in row:
				mev.append(cell.value)
		self.mev = np.array(mev, dtype=float)

		# load the remaining data, which are coefficients
		cs = []
		for row in sheet.iter_rows(min_row=2, min_col=2, max_col=len(header)):
			c = []
			for cell in row:
				c.append(cell.value)
			cs.append(c)
		self.coeffs = np.array(cs, dtype=float).transpose()


	def coeff(self, input):
		"""Given a material name, this returns the coeff for that material"""

		# check the material exists
		if input not in self.name:
			raise IndexError('Material ' + input + ' not found. Acceptable materials include: ' + str(self.name))

		# return the appropriate coeff
		index = self.name.index(input)
		return self.coeffs[index]